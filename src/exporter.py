import uuid
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


EXPORTS_DIR = Path("exports")
EXPORTS_DIR.mkdir(exist_ok=True)


def generate_xlsx(rows: list[dict], title: str = "Relatório") -> str:
    """Gera um arquivo XLSX a partir de uma lista de dicts e retorna o file_id."""
    if not rows:
        raise ValueError("Nenhum dado para exportar.")

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limita a 31 chars

    headers = list(rows[0].keys())

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows[:100]:  # sample first 100 rows
            val = str(row.get(header, ""))
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Save
    file_id = str(uuid.uuid4())
    filepath = EXPORTS_DIR / f"{file_id}.xlsx"
    wb.save(filepath)

    return file_id


def get_export_path(file_id: str) -> Path | None:
    filepath = EXPORTS_DIR / f"{file_id}.xlsx"
    if filepath.exists():
        return filepath
    return None


def cleanup_old_exports(max_age_hours: int = 24):
    """Remove arquivos de exportação com mais de max_age_hours."""
    now = datetime.now().timestamp()
    for f in EXPORTS_DIR.glob("*.xlsx"):
        age_hours = (now - f.stat().st_mtime) / 3600
        if age_hours > max_age_hours:
            f.unlink()
